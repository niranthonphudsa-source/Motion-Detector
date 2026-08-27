import os
import shutil
import time
import cv2
import numpy as np
from app.app import TableViewerWindow, ConfigManager
import json
import threading
import run_start.default_config_var as df
import csv
import shutil
from datetime import datetime
from openpyxl import Workbook, load_workbook
from datetime import datetime


def load_data():
    if os.path.exists('db_config.json'):
        try:
            with open('db_config.json', "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None
    else:
        print("File Not Found")
        return None
config_data = load_data()

class UserStateManager:

    def __init__(
        self,
        check_pose,
        fourcc,
        ok_display_time,
        max_lost_time,
        max_distance,
        buffer_output_time=5,
        save_ok=True,
        save_ng=True,
    ):
        self.user_states = {}
        self.check_pose = check_pose
        self.fourcc = fourcc
        self.ok_display_time = ok_display_time
        self.max_lost_time = max_lost_time
        self.max_distance = max_distance
        self.buffer_output_time = buffer_output_time
        self.save_ng = save_ng  # 🌟 รับค่าเริ่มต้นการเซฟ NG
        self.save_ok = save_ok  # 🌟 รับค่าเริ่มต้นการเซฟ OK
        self.user_id = None
        self.camera_id = None
        self.status_pose = None

    def get_or_recover_id(self, current_id, current_frame_active_ids, point_pose):
        if len(point_pose) < 17:
            return None

        current_time = time.time()
        curr_x, curr_y = int(point_pose[16][0]), int(point_pose[16][1])

        # 1. เช็กว่า ID นี้มีอยู่แล้วในระบบหรือไม่
        if current_id in self.user_states:
            return self.user_states[current_id]

        # 2. กรณีเป็น ID ใหม่ -> ลองค้นหา ID เก่าที่หายไปเพื่อสวมรอย (Recover)
        reclaimed_id = None
        for old_id, old_state in self.user_states.items():
            if old_id not in current_frame_active_ids:
                if old_state.get("last_seen_time") is not None:
                    time_diff = current_time - old_state["last_seen_time"]
                    if (
                        time_diff < self.max_lost_time
                        and old_state["was_inside_last_frame"]
                    ):
                        if old_state.get("last_position") is not None:
                            old_x, old_y = old_state["last_position"]
                            distance = np.sqrt(
                                (curr_x - old_x) ** 2 + (curr_y - old_y) ** 2
                            )
                            if distance < self.max_distance:
                                reclaimed_id = old_id
                                break

        if reclaimed_id is not None:
            self.user_states[current_id] = self.user_states.pop(reclaimed_id)
            self.user_states[current_id]["is_terminating"] = False
            self.user_states[current_id]["termination_start_time"] = None
            print(
                f"🔄 [ID Recovered] กู้คืนข้อมูลสำเร็จ: ID {reclaimed_id} -> ID {current_id}"
            )
        else:
            self.user_states[current_id] = {
                "valaus_last": [],
                "confirm": "NG",
                "is_ok_holding": False,
                "ok_start_time": 0,
                "video_filename": None,
                "writer": None,
                "was_inside_last_frame": False,
                "last_seen_time": None,
                "last_position": None,
                "is_terminating": False,
                "termination_start_time": None,
                "last_logged_sec": -1,
            }
        return self.user_states[current_id]

    def update_user_video(
        self, user_id, frame, is_inside_roi, save_ok=None, save_ng=None
    ):
        """🎬 ฟังก์ชันจัดการการเขียนวิดีโอ"""
        # 🌟 อัปเดตสถานะ save_ok / save_ng (ถ้ามีส่งเข้ามาระหว่างรัน)
        if save_ok is not None:
            self.save_ok = save_ok
        if save_ng is not None:
            self.save_ng = save_ng

        # 🌟 [เพิ่มจุดเช็ก] ถ้าทั้ง save_ok และ save_ng เป็น False ทั้งคู่ ไม่ต้องอัดวิดีโอใดๆ เลย
        if not self.save_ok and not self.save_ng:
            return

        state = self.user_states.get(user_id)
        if not state:
            return

        current_time = time.time()
        height, width = frame.shape[:2]

        # ─── 1. จังหวะอยู่ในจุดเช็ก (ROI = True) ───
        if is_inside_roi:
            if state["writer"] is None:
                os.makedirs("temp_video", exist_ok=True)
                timestamp = time.strftime("%Y%m%d_%H%M%S")
                filename = os.path.join(
                    "temp_video", f"user_{user_id}_{timestamp}.mp4"
                )

                state["writer"] = cv2.VideoWriter(
                    filename, self.fourcc, df.save_video_per_frame, (width, height)
                )
                state["video_filename"] = filename
                print(
                    f"🎥 [Start Recording] ID {user_id} เข้าจุดเช็ก เริ่มบันทึกวิดีโอ -> {filename}"
                )

            if state["is_terminating"]:
                state["is_terminating"] = False
                state["termination_start_time"] = None
                print(
                    f"⏩ ID {user_id} กลับเข้ามาในจุดเช็กอีกครั้ง ยกเลิกการนับถอยหลัง 5 วินาที"
                )

        # ─── 2. จังหวะก้าวออกจากจุดเช็ก (ROI = False) ───
        else:
            if state["writer"] is not None and not state["is_terminating"]:
                state["is_terminating"] = True
                state["termination_start_time"] = current_time
                print(
                    f"⏱️ ID {user_id} ออกจากจุดเช็ก เริ่มนับถอยหลังบันทึกแถมอีก {self.buffer_output_time} วินาที"
                )

        # ─── 3. บันทึกภาพลงไฟล์ ───
        if state["writer"] is not None:
            state["writer"].write(frame)

    def update_tracking_data(self, state, is_inside, point_pose):
        """อัปเดตสถานะและพิกัดล่าสุดของบุคคล"""
        state["was_inside_last_frame"] = is_inside
        state["last_seen_time"] = time.time()

        valid_pts = [pt for pt in point_pose if pt[0] > 0 and pt[1] > 0]
        if valid_pts:
            avg_x = sum(p[0] for p in valid_pts) / len(valid_pts)
            avg_y = sum(p[1] for p in valid_pts) / len(valid_pts)
            state["last_position"] = (avg_x, avg_y)

    def handle_lost_people(
        self,
        current_frame_active_ids,
        save_ok,
        save_ng,
        save_data,
        stats_db=None,
        camera_id="Camera_1",
    ):
        self.save_ok = save_ok
        self.save_ng = save_ng
        self.save_data = save_data
        current_time = time.time()

        for active_id, active_state in list(self.user_states.items()):
            
            # ตรวจสอบคนที่อยู่ในช่วงนับถอยหลังปิดไฟล์
            if (
                active_state["is_terminating"]
                and active_state["termination_start_time"] is not None
            ):
                elapsed_time = (
                    current_time - active_state["termination_start_time"]
                )
                remaining = max(0.0, self.buffer_output_time - elapsed_time)

                current_sec = int(elapsed_time)
                if current_sec != active_state.get("last_logged_sec", -1):
                    active_state["last_logged_sec"] = current_sec
                    print(
                        f"⏳ ID {active_id} กำลังอัดวิดีโอแถมท้าย.. เหลืออีก {remaining:.1f} วินาที"
                    )

                # 🏁 เมื่อครบเวลาบันทึกแถมพอดี -> ปิดวิดีโอ ย้ายไฟล์ และบันทึก Log
                if elapsed_time >= self.buffer_output_time:
                    # 1. คืนทรัพยากร VideoWriter
                    if active_state["writer"] is not None:
                        active_state["writer"].release()
                        active_state["writer"] = None
                        print(
                            f"🛑 [Stop Recording] บันทึกแถมครบ {self.buffer_output_time} วินาทีแล้ว ปิดไฟล์วิดีโอ ID {active_id}"
                        )

                    # ตัวอย่างการใช้งานตอนบันทึกไฟล์วิดีโอเสร็จ
                    # log_video_csv(active_camera_id, "OK", "video_ok/cam1_20260810_113000.mp4")

                    # 3. ตรวจสอบเงื่อนไขการย้าย/คัดลอกไฟล์วิดีโอ (ทำงานเฉพาะเมื่อมีไฟล์วิดีโอถูกสร้างขึ้นมาเท่านั้น)
                    is_ok = active_state["confirm"] == "OK"
                    should_save = self.save_ok if is_ok else self.save_ng
                    temp_file = active_state["video_filename"]

                    if temp_file and os.path.exists(temp_file):
                        base_filename = os.path.basename(temp_file)
                        status = "OK" if is_ok else "NG"
                        dest_folder = "video_ok" if is_ok else "video_ng"

                        os.makedirs(dest_folder, exist_ok=True)
                        dest_path = os.path.join(dest_folder, base_filename)

                        if should_save:
                            try:
                                shutil.move(temp_file, dest_path)
                                print(f"📁 [SUCCESS] ย้ายไฟล์วิดีโอสำเร็จไปที่: {dest_path}")

                                # ─── บันทึก Log ลงไฟล์ Excel (.xlsx) ───
                                log_file = "logs/video_history.xlsx"#..\logs\video_history.xlsx
                                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                cam_id = active_state.get("cam_id", camera_id)

                                # ตรวจสอบว่ามีไฟล์ Excel อยู่แล้วหรือไม่
                                if not os.path.exists(log_file):
                                    wb = Workbook()
                                    ws = wb.active
                                    ws.title = "Video Log"
                                    # สร้าง Header สำหรับไฟล์ใหม่
                                    ws.append(["Timestamp", "Camera_ID", "Status", "Video_Path"])
                                else:
                                    wb = load_workbook(log_file)
                                    ws = wb.active

                                # เพิ่มข้อมูลแถวใหม่ต่อท้ายไฟล์
                                ws.append([timestamp, cam_id, status, dest_path])
                                
                                # บันทึกไฟล์ Excel
                                wb.save(log_file)

                            except Exception as e:
                                print(f"❌ [ERROR] ไม่สามารถย้ายไฟล์หรือบันทึก Log ได้: {e}")
                        else:
                            try:
                                os.remove(temp_file)
                                print(
                                    f"🗑️ [CLEANUP] ลบไฟล์ชั่วคราวเรียบร้อย: {temp_file}"
                                )
                            except Exception as e:
                                print(
                                    f"❌ [ERROR] ลบไฟล์ชั่วคราวไม่สำเร็จ: {e}"
                                )

                if self.save_data:
                    elapsed_time = (
                        current_time - active_state["termination_start_time"]
                    )
                    remaining = max(0.0, self.buffer_output_time - elapsed_time)

                    current_sec = int(elapsed_time)
                    if current_sec != active_state.get("last_logged_sec", -1):
                        active_state["last_logged_sec"] = current_sec
                        print(
                            f"⏳ ID {active_id} กำลังนับถอยหลัง.. เหลืออีก {remaining:.1f} วินาที"
                        )

                    # 🏁 เมื่อครบเวลาบันทึกแถมพอดี -> ปิดวิดีโอ ย้ายไฟล์ และบันทึก Log
                    if elapsed_time >= self.buffer_output_time:
                        # 1. คืนทรัพยากร VideoWriter
                        if active_state["writer"] is not None:
                            active_state["writer"].release()
                            active_state["writer"] = None
                            print(
                                f"🛑 นับถอยหลังครบ {self.buffer_output_time} วินาทีแล้ว  ID {active_id}"
                            )
                    final_status = active_state["confirm"]
                    data = (active_id, camera_id, final_status)

                    def safe_insert_data(cfg, *d_args):
                        try:
                            
                            TableViewerWindow.insert_data(cfg, *d_args, self.save_data)
                        except Exception as e:
                            print(
                                f"⚠️ [DB Insert Error] ไม่สามารถเพิ่มข้อมูลลง TableViewer ได้: {e}"
                            )
                    db_thread = threading.Thread(
                        target=safe_insert_data,
                        args=(config_data, *data),
                        daemon=True,
                    )
                    db_thread.start()
                    # 4. Reset ค่าเพื่อเตรียมรับการทำงานรอบใหม่
                    active_state["video_filename"] = None
                    active_state["is_terminating"] = False
                    active_state["termination_start_time"] = None
                    active_state["last_logged_sec"] = -1
                    if active_state["confirm"] != "OK":
                        active_state["valaus_last"] = []

    def close_all_writers(self):
        """สั่งปิดฟังก์ชันการเขียนไฟล์ทั้งหมดเมื่อกดปิดโปรแกรม"""
        for active_id, active_state in self.user_states.items():
            if active_state["writer"] is not None:
                active_state["writer"].release()